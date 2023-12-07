import sys
import openpyxl
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from Scripts.Execution.extraction import openingTheBrowserAndExecutingPreconditions
from Scripts.Execution.generate_xpath import Xpath_Util
from UserLibraries.masterTemplateWriter import create_master_template_xpath
from UserLibraries.utilities import execute_preconditions
from selenium.webdriver.chrome.service import Service
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

def scanxpath(df,service,options):
    xpath_obj = Xpath_Util()
    updated_xpath=[]
    bToggle = False
    pages = df['PAGE_TITLE'].unique()

    for page in pages:
        df1 = df[(df["PAGE_TITLE"] == page) & (df["VALIDATION"] != "No")]
        if len(df1) > 0:
            # To initiate new driver for_new_page
            if str(df1["PRECONDITIONS"].unique()[0]) != 'nan':
                page_list = df1["PRECONDITIONS"].unique()[0].split(":")
                driver = ""
                driver, url = openingTheBrowserAndExecutingPreconditions(driver, page_list, df, service, options)
                title1 = driver.title
                url1 = driver.current_url
            else:
                title1 = page
                url1 = df1['URL'].unique()[0]
                driver = webdriver.Chrome(service - service, options=options)
                driver.get(url1)
                if page == "FinnOne SS0":
                    try:
                        WebDriverWait(driver, 3).until(EC.alert_is_present(),
                                                       'Timed out waiting for PA creation' +
                                                       'confirmation popup to appear.')

                        alert = driver.switch_to.alert
                        alert.accept()
                        print("alert accepted")
                    except TimeoutException:
                        print("no alert")

                    # aTab = driver.find_element(By, XPATH, '//a')
                    # aTab.click()
                driver.maximize_window()

            page1 = driver.execute_script("return document.body.innerHTML").\
                    encode('utf-8').decode('latin-1')  # returns the inner HTML as a string
            soup1 = BeautifulSoup(page1, 'html.parser')
            print(f"Scanning started for PAGE_TITLE: {title1} and URL: {url1}")

            result_flag = xpath_obj.generate_xpath(soup1, url1, title1, driver)
            xpath_new_list = result_flag[1]
            # bToggle = False
            xpath_old_list = []
            xpath_old_list = []
            dfrow = df[(df["PAGE_TITLE"] == page) & (df["VALIDATION"] != "No")]

            for index, row in dfrow.iterrows():
                xpath_old_list.append(row['FIELD_XPATH'])
                if len(xpath_old_list) == len(xpath_new_list):
                    for i, xpath_old in enumerate(xpath_old_list):
                        if xpath_old == xpath_new_list[i]:
                            continue
                        else:
                            # df_dash = pd.Dataframe ()
                            strDefGuidePath = str(ROOT_DIR) + f'/Output/ui_execution_guide.xlsx'
                            df_page_data = pd.read_excel(strDefGuidePath, sheet_name='PAGE_DATA')
                            wb = load_workbook(strDefGuidePath, read_only=True)

                        if 'DashBoard' not in wb.sheetnames:

                            #df_dash, insert(0, 'PAGE TITLE, '')
                            # df-dash.insert(1, 'XPATH_OLO, '')
                            # df_dash.insert(2, 'XPATH_NEW, '')

                            master_dict = {'PAGE TITLE':[], 'XPATH_OLD':[], 'XPATH_NEW':[] }
                            master_template_df = pd.DataFrame(master_dict)

                            df_page = pd.Series(page)
                            df_xpath_old = pd.Series(xpath_old)
                            df_xpath_new = pd.Series(xpath_new_list[i])

                            master_template_df['PAGE TITLE'] = df_page.values
                            master_template_df['XPATH_OLD'] = df_xpath_old.values
                            master_template_df['XPATH_NEW']= df_xpath_new.values
                            # master template_df.to_excel(strDefGuidePath, sheet_name='DashBoard,index=False)

                            with pd.ExcelWriter(strDefGuidePath, engine='openpyxl') as writer:
                                df_page_data.to_excel(writer, 'PAGE_DATA', index = False)
                                master_template_df.to_excel(writer, 'DashBoard', index=False)
                            updated_xpath.append(xpath_old)
                            print("DASHBOARD CREATED")

                        else:
                            # df-dash = pd.read_excel(strDefGuidePath,. sheet_name="DashBoard")
                            master_template_df = pd.read_excel(strDefGuidePath, sheet_name="DashBoard")
                            xpath_list = master_template_df['XPATH_OLD'].unique()

                            # master_dict = ('PAGE TITLE': [], 'XPATH_OLD': [), 'XPATH.NEW': [])
                            # master_template_df = pd.Datafname(master.dict)

                            df_page = pd.Series(page)
                            df_xpath_old = pd.Series(xpath_old)
                            df_xpath_new = pd.Series(xpath_new_list[i])

                            if len(xpath_list) > 0:
                                # wb = _ openpyxl.load_workbook(strDefGuidePath)
                                # sheet =_wb.active
                                # last_row = sheet.max_row
                                # # print(last_row)
                                # if last_row> 2:
                                # sdf = pd.DataFrame()

                                sdf_dict = {'PAGE TITLE':[], 'XPATH_OLD':[], 'XPATH_NEW':[]}
                                sdf_template_df = pd.DataFrame(sdf_dict)

                                sdf_template_df['PAGE TITLE'] = df_page.values
                                sdf_template_df['XPATH_OLD'] = df_xpath_old.values
                                sdf_template_df['XPATH_NEW'] = df_xpath_new.values

                                combine = pd.concat([master_template_df, sdf_template_df], ignore_index=True)

                                # master_template_df['PAGE TITLE'] '= df_page.values
                                # master_template_df['XPATH_OLD'] = df_xpath_old.values
                                # master_template_df['XPATH_NEW'] = df-xpath_new.values
                                # df_dash.append(master_template_df)
                                # master-template.df.to_excel(strDefGuidefath. sheet_name='DashBoard',index=False)

                                with pd.ExcelWriter(strDefGuidePath, engine='openpyxl') as writer:
                                    df_page_data.to_excel(writer, 'PAGE_DATA', index=False)
                                combine.to_excel(writer, 'DashBoard', index=False)
                                print("DASHBOARD CREATED")

                        print(f"-(xpath_old) changed to (xpath_new_list[i]l")
                        df.loc[df['FIELD_XPATH'] == xpath_old, 'FIELD_XPATH'] = [xpath_new_list[i]]
                        bToggle = True
                        updated_xpath.append(xpath_old)

                elif len(xpath_old_list) < len(xpath_new_list):
                    for i, xpath in enumerate(xpath_old_list):
                        if xpath in xpath_new_list:
                            continue
                        else:
                            df.loc[df['FIELD_XPATH'] == xpath_old, 'FIELD_XPATH'] = [xpath_new_list[i]]
                            print(f" (xpath_old) changed to fxpath_new_list[i])")
                            bToggle = True
                            updated_xpath.append(xpath_old)
                else:
                    print(f"xpath count for page: {page} is not matching. Check if any element added or removed from this page")
            driver.close()
        else:
            continue
    return df, bToggle, updated_xpath

if __name__== '_main_':
    path = Path(__file__)
    ROOT_DIR = path.parent.parent.parent.absolute()
    driver_path = str(ROOT_DIR) + "\\Driver\\chromedriver.exe"
    manual_guide = str(ROOT_DIR) + "\\Input\\Manual_Guide\\master_databuilder.xlsx"
    service = Service(executable_path=driver_path)
    options = Options()
    # options = webdriver.ChromeOptions()
    # options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    df= pd.read_excel(manual_guide)
    df,bToggle, updated_xpath = scanxpath(df,service,options)

    if bToggle:
        df.to_excel(manual_guide, sheet_name="MasterData", index=False)
        print(f"master databuilder updated for xpath: {updated_xpath}")