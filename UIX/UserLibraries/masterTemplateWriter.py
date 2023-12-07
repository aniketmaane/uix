import os
import json
import sys
from pathlib import Path
import pandas as pd
#from Userlibaries.ConfigReader import config,ROOT_DIR
#from Userlibaries.raml_support import normalize_raml, Get_Xml_Scanner
#from Userlibaries.utilities import filecheck
import warnings
from openpyxl import load_workbook
warnings.filterwarnings('ignore')

#APP_ROOT = os.path.dirname(os.path.abspath(__file__))
path = Path(__file__)
ROOT_DIR = path.parent.parent.absolute()
manual_guide_path = str(ROOT_DIR) + "/Input/Manual_Guide"
master_databuilder = os.path.join(manual_guide_path, "master_databuilder.xlsx")
vOutcome = "FAIL:Default-->create_master_template"

def create_master_template_xpath(xpath_list, url_list, title_list, action_list):
    try:
        df_pre = pd.read_excel(master_databuilder)
        id_list = []
        for i in range(1, len(xpath_list) + 1):
            id_list.append(i)

        master_dict = {'APP': [], 'PAGE_TITLE': [], 'FIELD_XPATH': [], 'DATATYPE': [],
                       'MIN': [], 'MAX': [], 'URL': [], 'VALUE': [], 'ACTION': [],
                       'VALIDATION': [], 'PRECONDITIONS': [], 'WAIT_TIME': []}

        df_master_template = pd.DataFrame(master_dict)

        df_id = pd.Series(id_list)
        df_xpath = pd.Series(xpath_list)
        df_url = pd.Series(url_list)
        df_title = pd.Series(title_list)
        #df_type = pd.Series(type_list)
        df_action = pd.Series(action_list)
        #df_master_template['App'] = df_id.values
        df_master_template['FIELD_XPATH'] = df_xpath.values
        df_master_template['URL'] = df_url.values
        df_master_template['PAGE_TITLE'] = df_title.values
        #df_master_template['FIELD_TYPE'] = df_type.values
        df_master_template['ACTION'] = df_action.values

        df = pd.concat([df_pre, df_master_template])

        df.to_excel(master_databuilder, sheet_name="MasterData", index=False)

        # with pd.ExcelWriter(master_databuilder, engine='openpyxl', mode='a') as writer:
        #     df_master_template.to_excel(writer, sheet_name='MasterData', startrow=writer.sheets['MasterData'].max_row, index=False)

        vOutcome = "master_databuilder created successfully"

        # Keyword Guide Writer
        # wb = load_workbook(manual_guide_path, read_only=True)
        # if "Keyword_Guide" in wb.sheetnames:
        #     bToggle = False
        # else:
        #     bToggle = True
        #     create_keyword_guide(bToggle)
    except:
        vOutcome = 'FAIL:create_master_template-->' + str(sys.exc_info())
    finally:
        print(vOutcome)


def create_keyword_guide():
    master_keyword_dict = {'KEYWORD': [], 'DESCRIPTION': [], 'REQUIRED_COLUMN': []}
    master_keyword_template_df = pd.DataFrame(master_keyword_dict)
    keyword_list = ["Enter", "EnterText", "Submit", "Click", "ClickOnElement", "Check",
                    "CheckTheCheckbox","Select", "SelectFromDropdown", "SwitchToNewWindow", "SwitchToFrame",
                    "SwitchToParentFrame","SwitchToDefaultContent"]

    description_list = ["Enter the value into the inputbox required for that inputbox",
                        "Enter the value into the input-box for the xPath mentioned into FIELD_XPATH column of master_databuilder",
                        "To click on submit/login button", "To click on the element/radio button",
                        "To click on the element for the xPath mentioned into FIELD_XPATH column of master_databuilder",
                        "TO check the checkbox",
                        "To check the checkbox for the xPath mentioned into FIELD_XPATH column of master_databuilder",
                        "To select the dropdown",
                        "To select the dropdown for the xPath and value mentioned into FIELD_XPATH and VALUE column of master_databuilder",
                        "Switch to the newly opened window",
                        "Switch to the frame mentioned in the VALUE column of master_databuilder",
                        "Switch back to the parent frame of the current frame",
                        "Switch back to the default content (first frame)"]

    required_column_list = ["VALUE", "VALUE,XPATH", "", "VALUE", "XPATH", "VALUE", "XPATH", "VALUE", "VALUE, XPATH", "", "VALUE","", ""]
    df_keyword = pd.Series(keyword_list)
    df_description = pd.Series(description_list)
    df_required = pd.Series(required_column_list)
    master_keyword_template_df['KEYWORD'] = df_keyword.values
    master_keyword_template_df['DESCRIPTION'] = df_description.values
    master_keyword_template_df['REQUIRED_COLUMN'] = df_required.values

    # Keyword guide
    with pd.ExcelWriter(manual_guide_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        master_keyword_template_df.to_excel(writer, sheet_name="Keyword_Guide", index=False)
    print("Keyword_Guide created successfully!")
